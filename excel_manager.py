"""
Manages reading and writing the project .xlsx workbook.

Sheet layout:
  Sheet 1: "Cover Sheet"  — project metadata and rack summary
  Sheet 2: "CLI Tool Help" — usage instructions for the command-line interface
  Sheet 3+: one sheet per rack

Cover Sheet cells:
  A2: Software Version
  B2: Controller Name
  C2: IO Network Card Name
  A5:A* / B5:B*: rack name / IO bit count (auto-populated)

Rack sheet columns (1-indexed):
  A: Module Type (dropdown)
  B: Module Slot Number
  C: PLC Routine Name
  D: I/O Bit
  E: I/O Buffer Tag Name
  F: I/O Buffer Tag Description
"""

import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from models import Bit, Module, Rack, Project, MODULE_TYPE_DROPDOWN, ALL_MODULE_TYPES, DIGITAL_TYPES, ANALOG_TYPES, SAFETY_TYPES, OTHER_TYPES, IO_FAMILY_POINT, IO_FAMILY_FLEX, IO_FAMILY_CLX

COVER_SHEET = "Cover Sheet"
CAD_SHEET   = "CAD_Descriptions"
HELP_SHEET  = "CLI Tool Help"

COL_MOD_TYPE = 1   # A
COL_SLOT     = 2   # B
COL_ROUTINE  = 3   # C
COL_BIT      = 4   # D
COL_TAG      = 5   # E
COL_DESC     = 6   # F

THIN = Side(style="thin")
BORDER_BOTTOM = Border(bottom=THIN)
HEADER_BORDER = Border(bottom=Side(style="medium"))


# ---------------------------------------------------------------------------
# Workbook creation
# ---------------------------------------------------------------------------

def create_workbook(path: str, software_version: str, controller_name: str, io_network_card: str,
                    project_number: str = "", project_description: str = "") -> None:
    wb = Workbook()

    # Sheet 1 — Cover Sheet
    ws_cover = wb.active
    ws_cover.title = COVER_SHEET
    _setup_cover_sheet(ws_cover, software_version, controller_name, io_network_card,
                       project_number, project_description)

    # Sheet 2 — CLI Tool Help
    ws_help = wb.create_sheet(HELP_SHEET)
    _setup_cli_help_sheet(ws_help)

    wb.save(path)


def _setup_cover_sheet(ws, software_version: str, controller_name: str, io_network_card: str,
                       project_number: str = "", project_description: str = "") -> None:
    ws["A1"] = "Software Version"
    ws["B1"] = "Controller Name"
    ws["C1"] = "IO Network Card Name"
    ws["D1"] = "Project Number"
    ws["E1"] = "Project Description"

    ws["A2"] = software_version
    ws["B2"] = controller_name
    ws["C2"] = io_network_card
    ws["D2"] = project_number
    ws["E2"] = project_description

    ws["A4"] = "Rack Name"
    ws["B4"] = "IO Bit Count"
    ws["C4"] = "IO Family"

    for cell in [ws["A4"], ws["B4"], ws["C4"]]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 40


def _setup_cli_help_sheet(ws) -> None:
    CLI_COMMANDS = [
        ("init",               "Create a new project workbook."),
        ("add-rack",           "Add a rack to the workbook."),
        ("rename-rack",        "Rename an existing rack."),
        ("remove-rack",        "Remove a rack sheet and its Cover Sheet entry."),
        ("add-module",         "Add modules to an existing rack."),
        ("fill-tags",          (
            "Auto-fill blank tag names in column E from the module type and routine name. "
            "Routine names must start with R#### (e.g. R4103) for the drawing number to appear "
            "in the tag. Existing values are never overwritten; rows without a module type are skipped."
        )),
        ("fill-descriptions",  "Fill blank tag descriptions in column F with 'spare'."),
        ("generate",           "Generate .l5x files from the workbook."),
        ("generate-cad",       "Generate a CAD description .xlsx from the workbook."),
        ("validate",           "Check the workbook for errors and warnings without generating files."),
        ("list",               "List all racks and modules in the workbook."),
    ]

    ws["A1"] = "Usage:  python io_buffer_tool.py <command>  [--workbook <path>]  [--output <dir>]"
    ws["A1"].font = Font(bold=True)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 18

    ws["A2"] = "Command"
    ws["B2"] = "Description"
    for cell in [ws["A2"], ws["B2"]]:
        cell.font = Font(bold=True)
        cell.border = HEADER_BORDER

    for i, (cmd, desc) in enumerate(CLI_COMMANDS, start=3):
        ws.cell(row=i, column=1, value=cmd)
        ws.cell(row=i, column=2, value=desc).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    ws.row_dimensions[8].height = 42   # fill-tags row needs extra height for wrapped text


# ---------------------------------------------------------------------------
# Add rack
# ---------------------------------------------------------------------------

def add_rack(path: str, rack_name: str, modules: list, io_family: str = IO_FAMILY_POINT) -> None:
    """
    modules: list of (num_bits: int,) for each module — slot numbers auto-assigned 1..N.
    Creates a new rack sheet and updates the Cover Sheet summary.
    """
    wb = load_workbook(path)

    if rack_name in wb.sheetnames:
        raise ValueError(f"Rack '{rack_name}' already exists in workbook.")

    ws = wb.create_sheet(rack_name)
    _write_rack_sheet(ws, modules)
    _append_cover_summary(wb[COVER_SHEET], rack_name, io_family)

    wb.save(path)


def _write_rack_sheet(ws, modules: list) -> None:
    """modules: list of int (bit counts per slot, in slot order)."""
    headers = ["Module Type", "Module Slot Number", "PLC Routine Name",
               "I/O Bit", "I/O Buffer Tag Name", "I/O Buffer Tag Description"]
    col_widths = [22, 27, 25, 11, 22, 28.5]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    # Module type dropdown validator (applied per module start row)
    dv = DataValidation(
        type="list",
        formula1=f'"{MODULE_TYPE_DROPDOWN}"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)

    current_row = 2
    for slot, num_bits in enumerate(modules, start=1):
        start_row = current_row
        end_row = current_row + num_bits - 1

        # Fill bit index rows
        for bit_idx in range(num_bits):
            row = current_row + bit_idx
            ws.cell(row=row, column=COL_BIT, value=bit_idx)

        # Slot number (merged across all bit rows)
        ws.cell(row=start_row, column=COL_SLOT, value=slot)

        # Routine name placeholder (merged)
        ws.cell(row=start_row, column=COL_ROUTINE, value="ENTER ROUTINE NAME HERE")

        # Apply dropdown validation to module type cell (top of merge)
        dv.add(ws.cell(row=start_row, column=COL_MOD_TYPE))

        # Merge columns A, B, C across all bit rows for this module
        if num_bits > 1:
            for col in [COL_MOD_TYPE, COL_SLOT, COL_ROUTINE]:
                ws.merge_cells(
                    start_row=start_row, start_column=col,
                    end_row=end_row, end_column=col
                )

        # Bottom border on last row of this module
        for col in range(1, 7):
            cell = ws.cell(row=end_row, column=col)
            cell.border = BORDER_BOTTOM

        # Center alignment for slot and bit columns
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=COL_SLOT).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=COL_BIT).alignment = Alignment(horizontal="center", vertical="center")

        # Center/wrap merged cells
        for col in [COL_MOD_TYPE, COL_SLOT, COL_ROUTINE]:
            ws.cell(row=start_row, column=col).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        current_row = end_row + 1

    # End sentinel
    ws.cell(row=current_row, column=COL_SLOT, value="End")



def _append_cover_summary(ws_cover, rack_name: str, io_family: str = IO_FAMILY_POINT) -> None:
    # Find next empty row starting at row 5
    row = 5
    while ws_cover.cell(row=row, column=COL_MOD_TYPE).value is not None:
        row += 1
    ws_cover.cell(row=row, column=COL_MOD_TYPE, value=rack_name)
    # Bit count will be filled at generate time; leave a placeholder for now
    ws_cover.cell(row=row, column=COL_SLOT, value=f"=COUNTA('{rack_name}'!E2:E5000)")
    ws_cover.cell(row=row, column=3, value=io_family)


# ---------------------------------------------------------------------------
# Add modules to existing rack
# ---------------------------------------------------------------------------

def add_modules_to_rack(path: str, rack_name: str, new_modules: list) -> None:
    """
    new_modules: list of int (bit counts), appended after existing modules.
    Removes the 'End' sentinel, appends new module rows, re-adds sentinel.
    """
    wb = load_workbook(path)
    if rack_name not in wb.sheetnames:
        raise ValueError(f"Rack '{rack_name}' not found in workbook.")

    ws = wb[rack_name]

    # Find and remove End sentinel, get next slot number
    end_row = None
    next_slot = 1
    for row in ws.iter_rows(min_row=2, max_col=COL_SLOT):
        cell = row[COL_SLOT - 1]
        if cell.value == "End":
            end_row = cell.row
            break
        if isinstance(cell.value, (int, float)):
            next_slot = int(cell.value) + 1

    if end_row is None:
        raise ValueError(f"Could not find 'End' sentinel in rack '{rack_name}'.")

    ws.cell(row=end_row, column=COL_SLOT, value=None)

    # Rebuild the validation and write new modules from end_row
    dv = DataValidation(
        type="list",
        formula1=f'"{MODULE_TYPE_DROPDOWN}"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)

    current_row = end_row
    for i, num_bits in enumerate(new_modules):
        slot = next_slot + i
        start_row = current_row
        end_row_mod = current_row + num_bits - 1

        for bit_idx in range(num_bits):
            ws.cell(row=current_row + bit_idx, column=COL_BIT, value=bit_idx)

        ws.cell(row=start_row, column=COL_SLOT, value=slot)
        ws.cell(row=start_row, column=COL_ROUTINE, value="ENTER ROUTINE NAME HERE")
        dv.add(ws.cell(row=start_row, column=COL_MOD_TYPE))

        if num_bits > 1:
            for col in [COL_MOD_TYPE, COL_SLOT, COL_ROUTINE]:
                ws.merge_cells(
                    start_row=start_row, start_column=col,
                    end_row=end_row_mod, end_column=col
                )

        for col in range(1, 7):
            ws.cell(row=end_row_mod, column=col).border = BORDER_BOTTOM

        for row in range(start_row, end_row_mod + 1):
            ws.cell(row=row, column=COL_SLOT).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=COL_BIT).alignment = Alignment(horizontal="center", vertical="center")

        for col in [COL_MOD_TYPE, COL_SLOT, COL_ROUTINE]:
            ws.cell(row=start_row, column=col).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        current_row = end_row_mod + 1

    ws.cell(row=current_row, column=COL_SLOT, value="End")
    wb.save(path)


# ---------------------------------------------------------------------------
# Read workbook → Project
# ---------------------------------------------------------------------------

def read_project(path: str) -> Project:
    wb = load_workbook(path, data_only=True)
    ws_cover = wb[COVER_SHEET]

    software_version    = str(ws_cover["A2"].value or "").strip()
    controller_name     = str(ws_cover["B2"].value or "").strip()
    io_network_card     = str(ws_cover["C2"].value or "").strip()
    project_number      = str(ws_cover["D2"].value or "").strip()
    project_description = str(ws_cover["E2"].value or "").strip()

    if not software_version or not controller_name or not io_network_card:
        raise ValueError(
            "Cover Sheet is missing Software Version (A2), Controller Name (B2), "
            "or IO Network Card Name (C2)."
        )

    # Build io_family map from cover sheet rack summary rows (A5+, C5+)
    valid_families = {IO_FAMILY_POINT, IO_FAMILY_FLEX, IO_FAMILY_CLX}
    family_map = {}
    for row in range(5, ws_cover.max_row + 1):
        rname = ws_cover.cell(row=row, column=1).value  # column A — rack name
        fam   = ws_cover.cell(row=row, column=3).value  # column C — IO family
        if rname and str(rname).strip():
            rname_str = str(rname).strip()
            if not fam or not str(fam).strip():
                raise ValueError(
                    f"Cover Sheet row {row}: IO Family is missing for rack '{rname_str}'. "
                    f"Must be one of: {', '.join(sorted(valid_families))}."
                )
            fam_str = str(fam).strip()
            if fam_str not in valid_families:
                raise ValueError(
                    f"Cover Sheet row {row}: IO Family '{fam_str}' for rack '{rname_str}' is not recognized. "
                    f"Must be one of: {', '.join(sorted(valid_families))}."
                )
            family_map[rname_str] = fam_str

    racks = []
    for ws in wb.worksheets:
        if ws.title in (COVER_SHEET, CAD_SHEET, HELP_SHEET):
            continue
        rack = _read_rack_sheet(ws)
        if ws.title not in family_map:
            raise ValueError(
                f"Rack sheet '{ws.title}' has no corresponding entry on the Cover Sheet. "
                f"Add it to the Cover Sheet with a valid IO Family "
                f"({', '.join(sorted(valid_families))})."
            )
        rack.io_family = family_map[ws.title]
        if rack.modules:
            racks.append(rack)

    # Cross-rack routine uniqueness (intra-rack duplicates are caught in _read_rack_sheet)
    all_routines: dict[str, str] = {}  # routine → rack name
    for rack in racks:
        for mod in rack.modules:
            if mod.routine:
                if mod.routine in all_routines:
                    raise ValueError(
                        f"PLC Routine Name '{mod.routine}' appears in both rack "
                        f"'{all_routines[mod.routine]}' and rack '{rack.name}'. "
                        f"Routine names must be unique across all racks."
                    )
                all_routines[mod.routine] = rack.name

    return Project(
        software_version=software_version,
        controller_name=controller_name,
        io_network_card=io_network_card,
        project_number=project_number,
        project_description=project_description,
        racks=racks,
    )


# ---------------------------------------------------------------------------
# Tag name generation
# ---------------------------------------------------------------------------

_TAG_PREFIX = {
    "Input":            ("DI",  "dot"),
    "Output":           ("DO",  "dot"),
    "Safety Input":     ("DIS", "dot"),
    "Safety Output":    ("DOS", "dot"),
    "Analog Input":     ("AI",  "bracket"),
    "Analog Output":    ("AO",  "bracket"),
    "Thermocouple/RTD": ("AI",  "bracket"),
}

_ROUTINE_RE = re.compile(r'^R(\d{4})')


def _generate_tag(mod_type: str, routine: str, bit_index: int) -> str:
    prefix, notation = _TAG_PREFIX.get(mod_type, ("??", "dot"))
    m = _ROUTINE_RE.match(routine)
    xxxx = m.group(1) if m else "XXXX"
    if notation == "bracket":
        return f"{prefix}_{xxxx}[{bit_index}]"
    return f"{prefix}_{xxxx}.{bit_index}"


def _check_all_routine_uniqueness(wb) -> None:
    """Raise ValueError if any PLC routine name appears more than once across all rack sheets."""
    seen: dict[str, tuple[str, int]] = {}  # routine → (rack_name, slot)

    for ws in wb.worksheets:
        if ws.title in (COVER_SHEET, CAD_SHEET, HELP_SHEET):
            continue

        merged_values: dict[tuple[int, int], object] = {}
        for merge in ws.merged_cells.ranges:
            top_left_val = ws.cell(merge.min_row, merge.min_col).value
            for row in range(merge.min_row, merge.max_row + 1):
                for col in range(merge.min_col, merge.max_col + 1):
                    merged_values[(row, col)] = top_left_val

        for row in range(2, ws.max_row + 1):
            slot_val = ws.cell(row=row, column=COL_SLOT).value  # raw — None for non-top-left merged cells
            if not isinstance(slot_val, (int, float)) or isinstance(slot_val, bool):
                continue
            slot = int(slot_val)

            routine_raw = merged_values.get((row, COL_ROUTINE), ws.cell(row=row, column=COL_ROUTINE).value)
            routine = str(routine_raw or "").strip()
            if not routine or routine == "ENTER ROUTINE NAME HERE":
                continue

            if routine in seen:
                first_rack, first_slot = seen[routine]
                if first_rack == ws.title:
                    raise ValueError(
                        f"Rack sheet '{ws.title}': PLC Routine Name '{routine}' is used by "
                        f"both slot {first_slot} and slot {slot}. Routine names must be unique."
                    )
                else:
                    raise ValueError(
                        f"PLC Routine Name '{routine}' appears in both rack '{first_rack}' "
                        f"(slot {first_slot}) and rack '{ws.title}' (slot {slot}). "
                        f"Routine names must be unique across all racks."
                    )
            seen[routine] = (ws.title, slot)


def fill_tags(path: str, rack_name: str) -> tuple[int, list[int]]:
    """
    Fill blank column-E cells with auto-generated tag names.
    Returns (filled_count, skipped_slots) where skipped_slots are slot numbers
    whose module type was not set.
    """
    wb = load_workbook(path)
    if rack_name not in wb.sheetnames:
        raise ValueError(f"Rack '{rack_name}' not found in workbook.")

    _check_all_routine_uniqueness(wb)

    ws = wb[rack_name]

    # Resolve merged cell values
    merged_values = {}
    for merge in ws.merged_cells.ranges:
        top_left_val = ws.cell(merge.min_row, merge.min_col).value
        for row in range(merge.min_row, merge.max_row + 1):
            for col in range(merge.min_col, merge.max_col + 1):
                merged_values[(row, col)] = top_left_val

    def cell_val(row, col):
        key = (row, col)
        if key in merged_values:
            return merged_values[key]
        return ws.cell(row=row, column=col).value

    filled = 0
    skipped_slots = []

    for row in range(2, ws.max_row + 1):
        bit_val = ws.cell(row=row, column=COL_BIT).value
        if not isinstance(bit_val, (int, float)) or isinstance(bit_val, bool):
            continue

        # Skip if tag already filled
        existing = ws.cell(row=row, column=COL_TAG).value
        if existing and str(existing).strip():
            continue

        mod_type = str(cell_val(row, COL_MOD_TYPE) or "").strip()
        if not mod_type or mod_type not in _TAG_PREFIX:
            # "Other" modules intentionally have no tags — don't report as skipped
            if mod_type not in OTHER_TYPES:
                slot = cell_val(row, COL_SLOT)
                slot_num = int(slot) if isinstance(slot, (int, float)) else None
                if slot_num is not None and slot_num not in skipped_slots:
                    skipped_slots.append(slot_num)
            continue

        routine = str(cell_val(row, COL_ROUTINE) or "").strip()
        tag = _generate_tag(mod_type, routine, int(bit_val))
        ws.cell(row=row, column=COL_TAG).value = tag
        filled += 1

    wb.save(path)
    return filled, skipped_slots


def rename_rack(path: str, old_name: str, new_name: str) -> None:
    """
    Rename a rack sheet and update the Cover Sheet summary row to match.
    Raises ValueError if old_name doesn't exist or new_name is already taken.
    """
    wb = load_workbook(path)

    if old_name not in wb.sheetnames:
        raise ValueError(f"Rack '{old_name}' not found in workbook.")
    if new_name in wb.sheetnames:
        raise ValueError(f"A sheet named '{new_name}' already exists in workbook.")

    # Rename the sheet
    wb[old_name].title = new_name

    # Update Cover Sheet: find the row where column A == old_name
    ws_cover = wb[COVER_SHEET]
    found = False
    for row in range(5, ws_cover.max_row + 1):
        cell_name = ws_cover.cell(row=row, column=COL_MOD_TYPE)
        if cell_name.value == old_name:
            cell_name.value = new_name
            # Rebuild the COUNTA formula with the new sheet name
            ws_cover.cell(row=row, column=COL_SLOT).value = f"=COUNTA('{new_name}'!E2:E5000)"
            found = True
            break

    if not found:
        raise ValueError(
            f"Rack '{old_name}' was not found in the Cover Sheet summary. "
            f"Sheet renamed, but Cover Sheet was not updated."
        )

    wb.save(path)


def remove_rack(path: str, rack_name: str) -> None:
    """
    Delete a rack sheet and its Cover Sheet summary row.
    Raises ValueError if rack_name doesn't exist.
    """
    wb = load_workbook(path)

    if rack_name not in wb.sheetnames:
        raise ValueError(f"Rack '{rack_name}' not found in workbook.")

    # Remove the rack sheet
    del wb[rack_name]

    # Remove the Cover Sheet summary row for this rack
    ws_cover = wb[COVER_SHEET]
    found = False
    for row in range(5, ws_cover.max_row + 1):
        cell_name = ws_cover.cell(row=row, column=COL_MOD_TYPE)
        if cell_name.value == rack_name:
            ws_cover.delete_rows(row)
            found = True
            break

    if not found:
        raise ValueError(
            f"Rack sheet '{rack_name}' was deleted, but no matching row was found on the Cover Sheet."
        )

    wb.save(path)


def fill_descriptions(path: str, rack_name: str) -> int:
    """
    Fill blank column-F (description) cells with 'spare'.
    Only touches rows that have a bit index in column D.
    Never overwrites existing values.
    Returns the count of cells filled.
    """
    wb = load_workbook(path)
    if rack_name not in wb.sheetnames:
        raise ValueError(f"Rack '{rack_name}' not found in workbook.")

    ws = wb[rack_name]
    filled = 0

    for row in range(2, ws.max_row + 1):
        bit_val = ws.cell(row=row, column=COL_BIT).value
        if not isinstance(bit_val, (int, float)) or isinstance(bit_val, bool):
            continue

        existing = ws.cell(row=row, column=COL_DESC).value
        if existing and str(existing).strip():
            continue

        ws.cell(row=row, column=COL_DESC).value = "spare"
        filled += 1

    wb.save(path)
    return filled


def _read_rack_sheet(ws) -> Rack:
    rack = Rack(name=ws.title)

    # Resolve merged cell values: openpyxl returns None for non-top-left merged cells.
    # Build a lookup of merged ranges so we can find the top-left value.
    merged_values = {}
    for merge in ws.merged_cells.ranges:
        top_left_val = ws.cell(merge.min_row, merge.min_col).value
        for row in range(merge.min_row, merge.max_row + 1):
            for col in range(merge.min_col, merge.max_col + 1):
                merged_values[(row, col)] = top_left_val

    def cell_val(row, col):
        key = (row, col)
        if key in merged_values:
            return merged_values[key]
        v = ws.cell(row=row, column=col).value
        return v

    # Identify module start rows using RAW cell values (not merged resolution).
    # Merged cells in col B have the slot number only in the top-left cell;
    # all other rows in the merge return None. Using cell_val() here would
    # incorrectly treat every merged row as a new module start.
    max_row = ws.max_row
    module_starts = []
    for row in range(2, max_row + 1):
        val = ws.cell(row=row, column=COL_SLOT).value  # raw — None for non-top-left merged cells
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            module_starts.append((row, int(val)))

    seen_routines = {}  # routine name → first slot number, for duplicate detection
    seen_tags = {}      # tag name → (slot, row), for duplicate detection across sheet
    for idx, (start_row, slot) in enumerate(module_starts):
        # Module ends one row before the next module start (or at max_row)
        if idx + 1 < len(module_starts):
            end_row = module_starts[idx + 1][0] - 1
        else:
            # Find End sentinel or next slot — again use raw values
            end_row = start_row
            for r in range(start_row + 1, max_row + 1):
                v = ws.cell(row=r, column=COL_SLOT).value  # raw
                if v == "End" or (isinstance(v, (int, float)) and not isinstance(v, bool)):
                    end_row = r - 1
                    break
                end_row = r

        mod_type = str(cell_val(start_row, COL_MOD_TYPE) or "").strip()
        routine  = str(cell_val(start_row, COL_ROUTINE) or "").strip()

        if routine in ("ENTER ROUTINE NAME HERE", ""):
            routine = ""

        if not routine:
            raise ValueError(
                f"Rack sheet '{ws.title}', slot {slot} (row {start_row}): "
                f"PLC Routine Name is missing."
            )
        if routine in seen_routines:
            raise ValueError(
                f"Rack sheet '{ws.title}', slot {slot} (row {start_row}): "
                f"PLC Routine Name '{routine}' is already used by slot {seen_routines[routine]}."
            )
        seen_routines[routine] = slot

        is_analog = mod_type in ANALOG_TYPES
        is_digital_or_safety = mod_type in (DIGITAL_TYPES | SAFETY_TYPES)
        is_other = mod_type in OTHER_TYPES

        bits = []
        if not is_other:
            for row in range(start_row, end_row + 1):
                bit_idx = ws.cell(row=row, column=COL_BIT).value
                if bit_idx is None:
                    continue
                tag  = str(ws.cell(row=row, column=COL_TAG).value or "").strip()
                desc = str(ws.cell(row=row, column=COL_DESC).value or "").strip()

                if not tag:
                    raise ValueError(
                        f"Rack sheet '{ws.title}', slot {slot}, row {row}: Tag name (column E) is missing."
                    )
                if tag in seen_tags:
                    first_slot, first_row = seen_tags[tag]
                    raise ValueError(
                        f"Rack sheet '{ws.title}', slot {slot}, row {row}: "
                        f"Tag '{tag}' is already used by slot {first_slot} (row {first_row})."
                    )
                if is_digital_or_safety and "." not in tag:
                    raise ValueError(
                        f"Rack sheet '{ws.title}', slot {slot}, row {row}: "
                        f"Tag '{tag}' is invalid for module type '{mod_type}' — expected a '.' (e.g. ROUTINE_NAME.0)."
                    )
                if is_analog and ("[" not in tag or "]" not in tag):
                    raise ValueError(
                        f"Rack sheet '{ws.title}', slot {slot}, row {row}: "
                        f"Tag '{tag}' is invalid for module type '{mod_type}' — expected '[]' (e.g. ROUTINE_NAME_AIN[0])."
                    )
                seen_tags[tag] = (slot, row)

                bits.append(Bit(index=int(bit_idx), tag=tag, description=desc))

        if not mod_type:
            raise ValueError(
                f"Rack sheet '{ws.title}', slot {slot} (row {start_row}): Module Type is blank. "
                f"Must be one of: {', '.join(ALL_MODULE_TYPES)}."
            )
        if mod_type not in ALL_MODULE_TYPES:
            raise ValueError(
                f"Rack sheet '{ws.title}', slot {slot} (row {start_row}): "
                f"Module Type '{mod_type}' is not recognized. "
                f"Must be one of: {', '.join(ALL_MODULE_TYPES)}."
            )
        rack.modules.append(Module(slot=slot, type=mod_type, routine=routine, bits=bits))

    return rack
