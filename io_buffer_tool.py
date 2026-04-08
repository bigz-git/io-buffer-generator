#!/usr/bin/env python3
"""
Point IO Buffer Generator — CLI

Usage:
  python tool.py init        Create a new project workbook
  python tool.py add-rack    Add a rack sheet to the workbook
  python tool.py add-module  Add modules to an existing rack
  python tool.py generate    Generate .l5x files from the workbook
  python tool.py list        List racks in the workbook
"""

import argparse
import glob
import os
import sys

import excel_manager
import l5x_generator
import cad_generator
from models import IO_FAMILY_POINT, IO_FAMILY_FLEX, IO_FAMILY_CLX


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_workbook(cwd: str) -> str | None:
    """Return the first .xlsx in cwd, or None."""
    matches = glob.glob(os.path.join(cwd, "*.xlsx"))
    return matches[0] if matches else None


def _prompt(prompt: str, default: str = "") -> str:
    suffix = f" [{default}]" if default else ""
    while True:
        val = input(f"{prompt}{suffix}: ").strip()
        if val:
            return val
        if default:
            return default
        print("  (required — please enter a value)")


def _prompt_int(prompt: str, min_val: int = 1) -> int:
    while True:
        raw = input(f"{prompt}: ").strip()
        try:
            val = int(raw)
            if val >= min_val:
                return val
            print(f"  Must be at least {min_val}.")
        except ValueError:
            print("  Please enter a whole number.")


def _get_workbook_path(args_path: str | None) -> str:
    if args_path:
        if not os.path.exists(args_path):
            print(f"Error: file not found: {args_path}")
            sys.exit(1)
        return args_path

    found = _find_workbook(os.getcwd())
    if found:
        print(f"Using workbook: {os.path.basename(found)}")
        return found

    print("No .xlsx workbook found in current directory.")
    print("Run 'python tool.py init' to create one, or use --workbook <path>.")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------

def cmd_init(args):
    os.system("")  # enable ANSI escape codes on Windows
    _BOLD  = "\033[1m"
    _CYAN  = "\033[94m"
    _WHITE = "\033[97m"
    _RESET = "\033[0m"
    print(f"\n{_BOLD}{_WHITE}━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"  {_CYAN}Quad Plus{_WHITE}  |  IO Buffer Generator Tool")
    print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━{_RESET}\n")

    filename = _prompt("Workbook filename (without .xlsx)", "project")
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    project_number      = _prompt("Project Number")
    project_description = _prompt("Project Description")
    out_dir = os.path.abspath(args.output) if args.output else os.getcwd()
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, filename)

    if os.path.exists(path):
        overwrite = input(f"'{filename}' already exists. Overwrite? [y/N]: ").strip().lower()
        if overwrite != "y":
            print("Aborted.")
            return

    print()
    software_version = _prompt("Software Version (e.g. 32.00)")
    controller_name  = _prompt("Controller Name")
    io_network_card  = _prompt("IO Network Card Name")

    excel_manager.create_workbook(path, software_version, controller_name, io_network_card,
                                  project_number, project_description)
    print(f"\nCreated: {filename}")


def cmd_add_rack(args):
    path = _get_workbook_path(args.workbook)

    print()
    rack_name = _prompt("Rack name")

    print("  IO Family:")
    print(f"    1. {IO_FAMILY_POINT} (Point IO)")
    print(f"    2. {IO_FAMILY_FLEX}  (Flex IO)")
    print(f"    3. {IO_FAMILY_CLX}   (ControlLogix IO)")
    while True:
        raw = input("  Select IO family [1]: ").strip()
        if not raw or raw == "1":
            io_family = IO_FAMILY_POINT
            break
        if raw == "2":
            io_family = IO_FAMILY_FLEX
            break
        if raw == "3":
            io_family = IO_FAMILY_CLX
            break
        print("  Please enter 1, 2, or 3.")

    num_modules = _prompt_int("Number of IO modules in this rack")

    print()
    modules = []
    for i in range(1, num_modules + 1):
        bits = _prompt_int(f"  Slot {i} — number of channels")
        modules.append(bits)

    try:
        excel_manager.add_rack(path, rack_name, modules, io_family)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)

    total_bits = sum(modules)
    print(f"\nAdded rack '{rack_name}' with {num_modules} module(s), {total_bits} total channels.")
    print(f"Open {os.path.basename(path)} to fill in tag names, descriptions, and routine names.")


def cmd_rename_rack(args):
    path = _get_workbook_path(args.workbook)

    from openpyxl import load_workbook as lw
    wb = lw(path, read_only=True)
    rack_names = [s for s in wb.sheetnames if s not in (excel_manager.COVER_SHEET, excel_manager.CAD_SHEET)]
    wb.close()

    if not rack_names:
        print("No racks found in workbook.")
        sys.exit(1)

    print()
    print("Available racks:")
    for i, name in enumerate(rack_names, 1):
        print(f"  {i}. {name}")

    while True:
        raw = input("Select rack to rename (number or name): ").strip()
        if raw.isdigit() and 1 <= int(raw) <= len(rack_names):
            old_name = rack_names[int(raw) - 1]
            break
        if raw in rack_names:
            old_name = raw
            break
        print("  Invalid selection.")

    new_name = _prompt("New rack name")

    try:
        excel_manager.rename_rack(path, old_name, new_name)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)

    print(f"\nRenamed '{old_name}' → '{new_name}'.")


def cmd_remove_rack(args):
    path = _get_workbook_path(args.workbook)

    from openpyxl import load_workbook as lw
    wb = lw(path, read_only=True)
    rack_names = [s for s in wb.sheetnames if s not in (excel_manager.COVER_SHEET, excel_manager.CAD_SHEET)]
    wb.close()

    if not rack_names:
        print("No racks found in workbook.")
        sys.exit(1)

    print()
    print("Available racks:")
    for i, name in enumerate(rack_names, 1):
        print(f"  {i}. {name}")

    while True:
        raw = input("Select rack to remove (number or name): ").strip()
        if raw.isdigit() and 1 <= int(raw) <= len(rack_names):
            rack_name = rack_names[int(raw) - 1]
            break
        if raw in rack_names:
            rack_name = raw
            break
        print("  Invalid selection.")

    confirm = input(f"Remove rack '{rack_name}'? This cannot be undone. [y/N]: ").strip().lower()
    if confirm != "y":
        print("Aborted.")
        return

    try:
        excel_manager.remove_rack(path, rack_name)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)

    print(f"\nRemoved rack '{rack_name}'.")


def cmd_add_module(args):
    path = _get_workbook_path(args.workbook)

    # Load workbook to show available racks
    from openpyxl import load_workbook as lw
    wb = lw(path, read_only=True)
    rack_names = [s for s in wb.sheetnames if s not in (excel_manager.COVER_SHEET, excel_manager.CAD_SHEET)]
    wb.close()

    if not rack_names:
        print("No racks found in workbook. Use 'add-rack' first.")
        sys.exit(1)

    print()
    print("Available racks:")
    for i, name in enumerate(rack_names, 1):
        print(f"  {i}. {name}")

    while True:
        raw = input("Select rack (number or name): ").strip()
        if raw.isdigit() and 1 <= int(raw) <= len(rack_names):
            rack_name = rack_names[int(raw) - 1]
            break
        if raw in rack_names:
            rack_name = raw
            break
        print("  Invalid selection.")

    num_modules = _prompt_int("Number of new modules to add")
    print()
    new_modules = []
    for i in range(1, num_modules + 1):
        bits = _prompt_int(f"  Module {i} — number of channels")
        new_modules.append(bits)

    try:
        excel_manager.add_modules_to_rack(path, rack_name, new_modules)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)

    print(f"\nAdded {num_modules} module(s) to rack '{rack_name}'.")


def cmd_fill_tags(args):
    path = _get_workbook_path(args.workbook)

    from openpyxl import load_workbook as lw
    wb = lw(path, read_only=True)
    rack_names = [s for s in wb.sheetnames if s not in (excel_manager.COVER_SHEET, excel_manager.CAD_SHEET)]
    wb.close()

    if not rack_names:
        print("No racks found. Use 'add-rack' first.")
        sys.exit(1)

    print()
    print("Available racks:")
    for i, name in enumerate(rack_names, 1):
        print(f"  {i}. {name}")
    print(f"  {len(rack_names) + 1}. All racks")

    while True:
        raw = input("Select rack (number or name): ").strip()
        if raw.isdigit():
            n = int(raw)
            if 1 <= n <= len(rack_names):
                selected = [rack_names[n - 1]]
                break
            if n == len(rack_names) + 1:
                selected = rack_names
                break
        else:
            match = next((n for n in rack_names if n.lower() == raw.lower()), None)
            if match:
                selected = [match]
                break
        print("  Invalid selection.")

    print()
    total_filled = 0
    for rack_name in selected:
        try:
            filled, skipped_slots = excel_manager.fill_tags(path, rack_name)
        except ValueError as e:
            print(f"Error: {e}")
            sys.exit(1)

        print(f"  {rack_name}: {filled} tag(s) filled.", end="")
        if skipped_slots:
            slots_str = ", ".join(str(s) for s in sorted(skipped_slots))
            print(f"  Warning: slot(s) {slots_str} skipped (module type not set).", end="")
        print()
        total_filled += filled

    print(f"\nDone. {total_filled} tag(s) written.")


def cmd_validate(args):
    path = _get_workbook_path(args.workbook)

    print("Validating workbook...")
    try:
        project = excel_manager.read_project(path)
    except ValueError as e:
        print(f"\nError: {e}")
        sys.exit(1)

    warnings = []

    for rack in project.racks:
        for mod in rack.modules:
            if not mod.routine:
                warnings.append(f"  Rack '{rack.name}', slot {mod.slot} ({mod.type}): no routine name.")
            missing_tags = sum(1 for b in mod.bits if not b.tag)
            if missing_tags:
                warnings.append(f"  Rack '{rack.name}', slot {mod.slot} ({mod.routine or '?'}): "
                                f"{missing_tags} of {len(mod.bits)} tag names missing.")
            missing_desc = sum(1 for b in mod.bits if not b.description)
            if missing_desc:
                warnings.append(f"  Rack '{rack.name}', slot {mod.slot} ({mod.routine or '?'}): "
                                f"{missing_desc} of {len(mod.bits)} descriptions missing.")

    if warnings:
        print(f"\nWarnings ({len(warnings)}):")
        for w in warnings:
            print(w)
    else:
        print("\nNo warnings.")

    total_tags  = sum(len(m.bits) for r in project.racks for m in r.modules)
    total_mods  = sum(len(r.modules) for r in project.racks)
    print(f"\nSummary: {len(project.racks)} rack(s), {total_mods} module(s), {total_tags} channel(s).")
    print("Workbook is valid." if not warnings else "Workbook passed structural checks but has warnings above.")


def cmd_fill_descriptions(args):
    path = _get_workbook_path(args.workbook)

    from openpyxl import load_workbook as lw
    wb = lw(path, read_only=True)
    rack_names = [s for s in wb.sheetnames if s not in (excel_manager.COVER_SHEET, excel_manager.CAD_SHEET)]
    wb.close()

    if not rack_names:
        print("No racks found. Use 'add-rack' first.")
        sys.exit(1)

    print()
    print("Available racks:")
    for i, name in enumerate(rack_names, 1):
        print(f"  {i}. {name}")
    print(f"  {len(rack_names) + 1}. All racks")

    while True:
        raw = input("Select rack (number or name): ").strip()
        if raw.isdigit():
            n = int(raw)
            if 1 <= n <= len(rack_names):
                selected = [rack_names[n - 1]]
                break
            if n == len(rack_names) + 1:
                selected = rack_names
                break
        elif raw in rack_names:
            selected = [raw]
            break
        print("  Invalid selection.")

    print()
    total_filled = 0
    for rack_name in selected:
        try:
            filled = excel_manager.fill_descriptions(path, rack_name)
        except ValueError as e:
            print(f"Error: {e}")
            sys.exit(1)
        print(f"  {rack_name}: {filled} description(s) filled.")
        total_filled += filled

    print(f"\nDone. {total_filled} description(s) written.")


def cmd_generate(args):
    path = _get_workbook_path(args.workbook)
    output_dir = os.path.abspath(args.output) if args.output else os.path.dirname(os.path.abspath(path))
    os.makedirs(output_dir, exist_ok=True)

    print("Reading workbook...")
    try:
        project = excel_manager.read_project(path)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)

    if not project.racks:
        print("No rack data found. Add racks and fill in tag information before generating.")
        sys.exit(1)

    # Warn about missing routine names
    missing = []
    for rack in project.racks:
        for mod in rack.modules:
            if not mod.routine:
                missing.append(f"  {rack.name} slot {mod.slot} ({mod.type})")
    if missing:
        print("Warning: the following modules have no routine name and will be skipped:")
        for m in missing:
            print(m)
        print()

    for rack in project.racks:
        for mod in rack.modules:
            missing_desc = sum(1 for b in mod.bits if not b.description)
            if missing_desc:
                print(f"Warning: Rack '{rack.name}', slot {mod.slot} ({mod.routine}): "
                      f"{missing_desc} of {len(mod.bits)} tag descriptions are missing.")

    print("Generating L5X files...")
    try:
        written = l5x_generator.generate(project, output_dir)
    except Exception as e:
        print(f"Error during generation: {e}")
        raise

    print()
    for f in written:
        print(f"  Written: {os.path.basename(f)}")
    print(f"\nDone. {len(written)} file(s) generated.")


def cmd_generate_cad(args):
    from datetime import datetime

    path = _get_workbook_path(args.workbook)
    output_dir = os.path.abspath(args.output) if args.output else os.path.dirname(os.path.abspath(path))
    os.makedirs(output_dir, exist_ok=True)

    # Prompt for module suffixes that cannot be extracted from the routine name.
    # The suffix is needed to determine the output row format for each module.
    missing = cad_generator.collect_missing_suffixes(path)
    suffix_overrides: dict[str, str] = {}
    if missing:
        print()
        print("The following routine names have no '_' separator — module type cannot be determined.")
        print("Enter the module type identifier for each (e.g. IA4, OW4, IB8):")
        for routine in missing:
            suffix_overrides[routine] = _prompt(f"  Module type for routine '{routine}'")

    filename = f"CAD_Descriptions_{datetime.now().strftime('%d%m%y_%H%M')}.xlsx"
    out_path = os.path.join(output_dir, filename)

    print(f"\nGenerating CAD description file...")
    cad_generator.generate_cad(path, out_path, suffix_overrides)
    print(f"Written: {filename}")


def cmd_list(args):
    path = _get_workbook_path(args.workbook)

    try:
        project = excel_manager.read_project(path)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)

    print(f"\nProject: {os.path.basename(path)}")
    print(f"  Software Version : {project.software_version}")
    print(f"  Controller       : {project.controller_name}")
    print(f"  IO Network Card  : {project.io_network_card}")
    print(f"\nRacks ({len(project.racks)}):")

    for rack in project.racks:
        total_bits = sum(len(m.bits) for m in rack.modules)
        print(f"\n  {rack.name}  ({len(rack.modules)} module(s), {total_bits} channels)  [{rack.io_family}]")
        for mod in rack.modules:
            routine = mod.routine or "(no routine name)"
            print(f"    Slot {mod.slot:>2}  {mod.type:<20}  {len(mod.bits):>2} channels  → {routine}")


# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Point IO Buffer Generator — generate Studio 5000 .l5x files from Excel."
    )
    parser.add_argument("--workbook", "-w", help="Path to project .xlsx file")
    parser.add_argument("--output", "-o", help="Target directory for generated files")

    sub = parser.add_subparsers(dest="command", metavar="command")
    sub.required = True

    sub.add_parser("validate",    help="Check workbook for errors and warnings without generating files")
    sub.add_parser("init",       help="Create a new project workbook")
    sub.add_parser("add-rack",    help="Add a rack to the workbook")
    sub.add_parser("rename-rack", help="Rename an existing rack")
    sub.add_parser("remove-rack", help="Remove a rack sheet and its Cover Sheet entry")
    sub.add_parser("add-module",  help="Add modules to an existing rack")
    sub.add_parser(
        "fill-tags",
        help="Auto-fill blank tag names in column E",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Tag names are generated from the module type (column A) and routine name (column C).\n"
            "Routine names must start with R#### for the drawing number to be used in the tag\n"
            "(e.g. R4103, where 4103 is the drawing sheet number where the I/O module is shown).\n"
            "Rows where column E is already filled are never overwritten.\n"
            "Rows where column A (module type) is not set are skipped with a warning."
        ),
    )
    sub.add_parser("fill-descriptions", help="Fill blank tag descriptions in column F with 'spare'")
    sub.add_parser("generate",     help="Generate .l5x files from the workbook")
    sub.add_parser("generate-cad", help="Generate CAD description .xlsx from the workbook")
    sub.add_parser("list",         help="List racks and modules in the workbook")

    args = parser.parse_args()

    commands = {
        "validate":          cmd_validate,
        "init":              cmd_init,
        "add-rack":          cmd_add_rack,
        "rename-rack":       cmd_rename_rack,
        "remove-rack":       cmd_remove_rack,
        "add-module":        cmd_add_module,
        "fill-tags":         cmd_fill_tags,
        "fill-descriptions": cmd_fill_descriptions,
        "generate":          cmd_generate,
        "generate-cad":      cmd_generate_cad,
        "list":              cmd_list,
    }
    commands[args.command](args)


if __name__ == "__main__":
    main()
