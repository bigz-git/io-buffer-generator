#!/usr/bin/env python3
"""
IO Buffer Generator — GUI
Wraps the CLI commands in a simple Tkinter launcher.
"""

import glob
import os
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import load_workbook
from PIL import Image, ImageTk

import excel_manager
import l5x_generator
from models import IO_FAMILY_CLX, IO_FAMILY_FLEX, IO_FAMILY_POINT

COVER_SHEET = excel_manager.COVER_SHEET
CAD_SHEET = excel_manager.CAD_SHEET


# ---------------------------------------------------------------------------
# Main window
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("IO Buffer Generator")
        self.minsize(720, 420)
        try:
            img = Image.open(os.path.join(os.path.dirname(__file__), "Quad Plus Brand Logo.png"))
            icons = [ImageTk.PhotoImage(img.resize((s, s), Image.LANCZOS)) for s in (16, 32, 48)]
            self.iconphoto(True, *icons)
            self._icons = icons  # prevent garbage collection
        except Exception:
            pass
        self._build_ui()
        self._auto_detect_workbook()

    def _build_ui(self):
        # ── Path selectors ──────────────────────────────────────────────────
        top = ttk.Frame(self, padding=8)
        top.pack(fill="x")

        ttk.Label(top, text="Workbook:").grid(row=0, column=0, sticky="w")
        self.wb_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.wb_var, width=55).grid(row=0, column=1, padx=4)
        ttk.Button(top, text="Browse…", command=self._browse_workbook).grid(row=0, column=2)
        ttk.Button(top, text="New…", command=self.cmd_init).grid(row=0, column=3, padx=(4, 0))

        ttk.Label(top, text="Output dir:").grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.out_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.out_var, width=55).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Button(top, text="Browse…", command=self._browse_output).grid(row=1, column=2, pady=(4, 0))
        ttk.Label(top, text="(defaults to workbook directory)", foreground="grey").grid(
            row=1, column=3, sticky="w", padx=(4, 0), pady=(4, 0))

        ttk.Separator(self, orient="horizontal").pack(fill="x", pady=(4, 0))

        # ── Body ────────────────────────────────────────────────────────────
        body = ttk.Frame(self, padding=(8, 6, 8, 8))
        body.pack(fill="both", expand=True)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        # Button column
        btn_col = ttk.Frame(body)
        btn_col.grid(row=0, column=0, sticky="ns", padx=(0, 10))

        def section(text):
            ttk.Label(btn_col, text=text, font=("", 9, "bold")).pack(anchor="w", pady=(10, 2))

        def btn(text, cmd):
            ttk.Button(btn_col, text=text, width=20, command=cmd).pack(fill="x", pady=1)

        section("Setup")
        btn("Add Rack",    self.cmd_add_rack)
        btn("Add Module",  self.cmd_add_module)
        btn("Rename Rack", self.cmd_rename_rack)
        btn("Remove Rack", self.cmd_remove_rack)

        section("Populate")
        btn("Fill Tags",         self.cmd_fill_tags)
        btn("Fill Descriptions", self.cmd_fill_descriptions)

        section("Tools")
        btn("List",     self.cmd_list)
        btn("Validate", self.cmd_validate)
        btn("Generate", self.cmd_generate)

        # Log area
        self.log = scrolledtext.ScrolledText(
            body, state="disabled", font=("Courier", 9), wrap="word", relief="sunken")
        self.log.grid(row=0, column=1, sticky="nsew")

    # ── Helpers ─────────────────────────────────────────────────────────────

    def _auto_detect_workbook(self):
        matches = glob.glob(os.path.join(os.getcwd(), "*.xlsx"))
        if matches:
            self.wb_var.set(matches[0])

    def _browse_workbook(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.wb_var.set(path)

    def _browse_output(self):
        path = filedialog.askdirectory()
        if path:
            self.out_var.set(path)

    def _workbook_path(self) -> str | None:
        path = self.wb_var.get().strip()
        if not path:
            messagebox.showerror("No Workbook", "Select or create a workbook first.")
            return None
        if not os.path.exists(path):
            messagebox.showerror("Not Found", f"File not found:\n{path}")
            return None
        return path

    def _get_rack_names(self, path: str) -> list[str]:
        wb = load_workbook(path, read_only=True)
        names = [s for s in wb.sheetnames if s not in (COVER_SHEET, CAD_SHEET)]
        wb.close()
        return names

    def _log(self, text: str):
        self.log.configure(state="normal")
        self.log.insert("end", text + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def _log_clear(self):
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")

    # ── Commands ─────────────────────────────────────────────────────────────

    def cmd_init(self):
        dlg = InitDialog(self)
        self.wait_window(dlg)
        if not dlg.result:
            return
        r = dlg.result
        filename = r["filename"]
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        out_dir = self.out_var.get().strip() or os.getcwd()
        path = os.path.join(out_dir, filename)
        if os.path.exists(path):
            if not messagebox.askyesno("Overwrite?", f"'{filename}' already exists. Overwrite?"):
                return
        try:
            excel_manager.create_workbook(
                path,
                r["software_version"],
                r["controller_name"],
                r["io_network_card"],
                r["project_number"],
                r["project_description"],
            )
            self.wb_var.set(path)
            self._log_clear()
            self._log(f"Created: {filename}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def cmd_add_rack(self):
        path = self._workbook_path()
        if not path:
            return
        dlg = AddRackDialog(self)
        self.wait_window(dlg)
        if not dlg.result:
            return
        r = dlg.result
        try:
            excel_manager.add_rack(path, r["name"], r["channels"], r["io_family"])
            self._log(f"Added rack '{r['name']}' — {len(r['channels'])} module(s), "
                      f"{sum(r['channels'])} total channels.")
        except ValueError as e:
            messagebox.showerror("Error", str(e))

    def cmd_add_module(self):
        path = self._workbook_path()
        if not path:
            return
        rack_names = self._get_rack_names(path)
        if not rack_names:
            messagebox.showerror("No Racks", "No racks found. Use 'Add Rack' first.")
            return
        dlg = AddModuleDialog(self, rack_names)
        self.wait_window(dlg)
        if not dlg.result:
            return
        r = dlg.result
        try:
            excel_manager.add_modules_to_rack(path, r["rack"], r["channels"])
            self._log(f"Added {len(r['channels'])} module(s) to rack '{r['rack']}'.")
        except ValueError as e:
            messagebox.showerror("Error", str(e))

    def cmd_rename_rack(self):
        path = self._workbook_path()
        if not path:
            return
        rack_names = self._get_rack_names(path)
        if not rack_names:
            messagebox.showerror("No Racks", "No racks found.")
            return
        dlg = RenameRackDialog(self, rack_names)
        self.wait_window(dlg)
        if not dlg.result:
            return
        old, new = dlg.result
        try:
            excel_manager.rename_rack(path, old, new)
            self._log(f"Renamed '{old}' → '{new}'.")
        except ValueError as e:
            messagebox.showerror("Error", str(e))

    def cmd_remove_rack(self):
        path = self._workbook_path()
        if not path:
            return
        rack_names = self._get_rack_names(path)
        if not rack_names:
            messagebox.showerror("No Racks", "No racks found.")
            return
        dlg = SelectRackDialog(self, rack_names, title="Remove Rack", label="Select rack to remove:")
        self.wait_window(dlg)
        if not dlg.result:
            return
        rack = dlg.result
        if not messagebox.askyesno("Confirm Remove",
                                   f"Remove rack '{rack}'?\nThis cannot be undone."):
            return
        try:
            excel_manager.remove_rack(path, rack)
            self._log(f"Removed rack '{rack}'.")
        except ValueError as e:
            messagebox.showerror("Error", str(e))

    def cmd_fill_tags(self):
        path = self._workbook_path()
        if not path:
            return
        rack_names = self._get_rack_names(path)
        if not rack_names:
            messagebox.showerror("No Racks", "No racks found.")
            return
        dlg = SelectRackDialog(self, rack_names, title="Fill Tags",
                               label="Select rack:", allow_all=True)
        self.wait_window(dlg)
        if not dlg.result:
            return
        selected = rack_names if dlg.result == "__ALL__" else [dlg.result]
        total = 0
        for rack_name in selected:
            try:
                filled, skipped = excel_manager.fill_tags(path, rack_name)
                msg = f"  {rack_name}: {filled} tag(s) filled."
                if skipped:
                    msg += f"  (slots skipped — type not set: {', '.join(str(s) for s in sorted(skipped))})"
                self._log(msg)
                total += filled
            except ValueError as e:
                messagebox.showerror("Error", str(e))
                return
        self._log(f"Done. {total} tag(s) written.")

    def cmd_fill_descriptions(self):
        path = self._workbook_path()
        if not path:
            return
        rack_names = self._get_rack_names(path)
        if not rack_names:
            messagebox.showerror("No Racks", "No racks found.")
            return
        dlg = SelectRackDialog(self, rack_names, title="Fill Descriptions",
                               label="Select rack:", allow_all=True)
        self.wait_window(dlg)
        if not dlg.result:
            return
        selected = rack_names if dlg.result == "__ALL__" else [dlg.result]
        total = 0
        for rack_name in selected:
            try:
                filled = excel_manager.fill_descriptions(path, rack_name)
                self._log(f"  {rack_name}: {filled} description(s) filled.")
                total += filled
            except ValueError as e:
                messagebox.showerror("Error", str(e))
                return
        self._log(f"Done. {total} description(s) written.")

    def cmd_validate(self):
        path = self._workbook_path()
        if not path:
            return
        self._log_clear()
        self._log("Validating workbook…")
        try:
            project = excel_manager.read_project(path)
        except ValueError as e:
            self._log(f"Error: {e}")
            return

        warnings = []
        for rack in project.racks:
            for mod in rack.modules:
                if not mod.routine:
                    warnings.append(
                        f"  Rack '{rack.name}', slot {mod.slot} ({mod.type}): no routine name.")
                missing_tags = sum(1 for b in mod.bits if not b.tag)
                if missing_tags:
                    warnings.append(
                        f"  Rack '{rack.name}', slot {mod.slot} ({mod.routine or '?'}): "
                        f"{missing_tags} of {len(mod.bits)} tag names missing.")
                missing_desc = sum(1 for b in mod.bits if not b.description)
                if missing_desc:
                    warnings.append(
                        f"  Rack '{rack.name}', slot {mod.slot} ({mod.routine or '?'}): "
                        f"{missing_desc} of {len(mod.bits)} descriptions missing.")

        if warnings:
            self._log(f"\nWarnings ({len(warnings)}):")
            for w in warnings:
                self._log(w)
        else:
            self._log("\nNo warnings.")

        total_tags = sum(len(m.bits) for r in project.racks for m in r.modules)
        total_mods = sum(len(r.modules) for r in project.racks)
        self._log(f"\nSummary: {len(project.racks)} rack(s), {total_mods} module(s), {total_tags} channel(s).")
        self._log("Workbook is valid." if not warnings
                  else "Workbook passed structural checks but has warnings above.")

    def cmd_list(self):
        path = self._workbook_path()
        if not path:
            return
        self._log_clear()
        try:
            project = excel_manager.read_project(path)
        except ValueError as e:
            self._log(f"Error: {e}")
            return

        self._log(f"Project: {os.path.basename(path)}")
        self._log(f"  Software Version : {project.software_version}")
        self._log(f"  Controller       : {project.controller_name}")
        self._log(f"  IO Network Card  : {project.io_network_card}")
        self._log(f"\nRacks ({len(project.racks)}):")
        for rack in project.racks:
            total_bits = sum(len(m.bits) for m in rack.modules)
            self._log(f"\n  {rack.name}  ({len(rack.modules)} module(s), {total_bits} channels)"
                      f"  [{rack.io_family}]")
            for mod in rack.modules:
                routine = mod.routine or "(no routine name)"
                self._log(f"    Slot {mod.slot:>2}  {mod.type:<20}  "
                          f"{len(mod.bits):>2} channels  → {routine}")

    def cmd_generate(self):
        path = self._workbook_path()
        if not path:
            return
        out_dir = self.out_var.get().strip() or os.path.dirname(os.path.abspath(path))
        os.makedirs(out_dir, exist_ok=True)

        self._log_clear()
        self._log("Reading workbook…")
        try:
            project = excel_manager.read_project(path)
        except ValueError as e:
            self._log(f"Error: {e}")
            return

        if not project.racks:
            self._log("No rack data found. Add racks and fill in tag information before generating.")
            return

        missing_routines = [
            f"  {rack.name} slot {mod.slot} ({mod.type})"
            for rack in project.racks
            for mod in rack.modules
            if not mod.routine
        ]
        if missing_routines:
            self._log("Warning: modules with no routine name will be skipped:")
            for m in missing_routines:
                self._log(m)
            self._log("")

        for rack in project.racks:
            for mod in rack.modules:
                missing_desc = sum(1 for b in mod.bits if not b.description)
                if missing_desc:
                    self._log(f"Warning: Rack '{rack.name}', slot {mod.slot} ({mod.routine}): "
                              f"{missing_desc} of {len(mod.bits)} tag descriptions are missing.")

        self._log("Generating L5X files…")
        try:
            written = l5x_generator.generate(project, out_dir)
        except Exception as e:
            self._log(f"Error during generation: {e}")
            return

        self._log("")
        for f in written:
            self._log(f"  Written: {os.path.basename(f)}")
        self._log(f"\nDone. {len(written)} file(s) generated in:\n  {out_dir}")


# ---------------------------------------------------------------------------
# Dialogs
# ---------------------------------------------------------------------------

class _BaseDialog(tk.Toplevel):
    def __init__(self, parent, title: str):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.result = None
        self._build()
        # Center over parent
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width()  - self.winfo_width())  // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

    def _build(self):
        raise NotImplementedError

    def _ok(self):
        raise NotImplementedError

    def _cancel(self):
        self.destroy()

    def _add_footer(self, parent):
        f = ttk.Frame(parent)
        f.pack(fill="x", pady=(8, 0))
        ttk.Button(f, text="Cancel", command=self._cancel).pack(side="right", padx=(4, 0))
        ttk.Button(f, text="OK", command=self._ok, default="active").pack(side="right")
        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self._cancel())


class InitDialog(_BaseDialog):
    _FIELDS = [
        ("filename",            "Filename (without .xlsx)",  "project"),
        ("project_number",      "Project Number",            ""),
        ("project_description", "Project Description",       ""),
        ("software_version",    "Software Version",          "32.00"),
        ("controller_name",     "Controller Name",           ""),
        ("io_network_card",     "IO Network Card Name",      ""),
    ]

    def __init__(self, parent):
        self._vars: dict[str, tk.StringVar] = {}
        super().__init__(parent, "New Project Workbook")

    def _build(self):
        outer = ttk.Frame(self, padding=14)
        outer.pack(fill="both")
        fields = ttk.Frame(outer)
        fields.pack(fill="x")
        for i, (key, label, default) in enumerate(self._FIELDS):
            ttk.Label(fields, text=label + ":").grid(row=i, column=0, sticky="w", pady=3, padx=(0, 8))
            var = tk.StringVar(value=default)
            self._vars[key] = var
            ttk.Entry(fields, textvariable=var, width=36).grid(row=i, column=1, pady=3)
        self._add_footer(outer)

    def _ok(self):
        vals = {k: v.get().strip() for k, v in self._vars.items()}
        if not vals["filename"]:
            messagebox.showerror("Required", "Filename is required.", parent=self)
            return
        self.result = vals
        self.destroy()


class _RackModuleBase(_BaseDialog):
    """Shared base for Add Rack and Add Module — both need a dynamic channel list."""

    def __init__(self, parent, title: str):
        self._channel_vars: list[tk.IntVar] = []
        super().__init__(parent, title)
        self.resizable(False, True)

    def _build_channel_area(self, parent, slot_label: str = "Slot") -> ttk.Frame:
        """Build the scrollable channel-count section. Returns the inner frame."""
        ttk.Label(parent, text="Channels per module:").pack(anchor="w", pady=(4, 2))
        container = ttk.Frame(parent)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, height=160, width=260, highlightthickness=0)
        sb = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        self._ch_inner = ttk.Frame(canvas)
        self._ch_win = canvas.create_window((0, 0), window=self._ch_inner, anchor="nw")
        self._ch_inner.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._canvas = canvas
        self._slot_label = slot_label
        return self._ch_inner

    def _rebuild_channels(self, n: int):
        for w in self._ch_inner.winfo_children():
            w.destroy()
        self._channel_vars = []
        n = max(1, min(n, 64))
        for i in range(n):
            ttk.Label(self._ch_inner, text=f"  {self._slot_label} {i + 1}:").grid(
                row=i, column=0, sticky="w", pady=1)
            var = tk.IntVar(value=16)
            ttk.Spinbox(self._ch_inner, from_=1, to=512, textvariable=var, width=7).grid(
                row=i, column=1, padx=8, pady=1)
            self._channel_vars.append(var)

    def _collect_channels(self) -> list[int] | None:
        try:
            ch = [v.get() for v in self._channel_vars]
        except tk.TclError:
            messagebox.showerror("Invalid", "Channel counts must be whole numbers.", parent=self)
            return None
        if any(c < 1 for c in ch):
            messagebox.showerror("Invalid", "Channel counts must be at least 1.", parent=self)
            return None
        return ch


class AddRackDialog(_RackModuleBase):
    def __init__(self, parent):
        super().__init__(parent, "Add Rack")

    def _build(self):
        outer = ttk.Frame(self, padding=14)
        outer.pack(fill="both", expand=True)

        top = ttk.Frame(outer)
        top.pack(fill="x")

        ttk.Label(top, text="Rack name:").grid(row=0, column=0, sticky="w", pady=3)
        self._name_var = tk.StringVar()
        ttk.Entry(top, textvariable=self._name_var, width=32).grid(
            row=0, column=1, padx=8, pady=3, sticky="w")

        ttk.Label(top, text="IO Family:").grid(row=1, column=0, sticky="nw", pady=3)
        self._family_var = tk.StringVar(value=IO_FAMILY_POINT)
        fam_frame = ttk.Frame(top)
        fam_frame.grid(row=1, column=1, sticky="w", padx=8)
        for val, label in [
            (IO_FAMILY_POINT, f"{IO_FAMILY_POINT} — Point IO"),
            (IO_FAMILY_FLEX,  f"{IO_FAMILY_FLEX}  — Flex IO"),
            (IO_FAMILY_CLX,   f"{IO_FAMILY_CLX}   — ControlLogix IO"),
        ]:
            ttk.Radiobutton(fam_frame, text=label, variable=self._family_var,
                            value=val).pack(anchor="w")

        ttk.Label(top, text="Number of modules:").grid(row=2, column=0, sticky="w", pady=3)
        self._num_var = tk.StringVar(value="1")
        ttk.Spinbox(top, from_=1, to=64, textvariable=self._num_var, width=7,
                    command=self._on_num_change).grid(row=2, column=1, sticky="w", padx=8, pady=3)
        self._num_var.trace_add("write", lambda *_: self.after_idle(self._on_num_change))

        ttk.Separator(outer, orient="horizontal").pack(fill="x", pady=6)
        self._build_channel_area(outer, slot_label="Slot")
        self._rebuild_channels(1)
        self._add_footer(outer)

    def _on_num_change(self):
        try:
            n = int(self._num_var.get())
        except ValueError:
            return
        self._rebuild_channels(n)

    def _ok(self):
        name = self._name_var.get().strip()
        if not name:
            messagebox.showerror("Required", "Rack name is required.", parent=self)
            return
        channels = self._collect_channels()
        if channels is None:
            return
        self.result = {"name": name, "io_family": self._family_var.get(), "channels": channels}
        self.destroy()


class AddModuleDialog(_RackModuleBase):
    def __init__(self, parent, rack_names: list[str]):
        self._rack_names = rack_names
        super().__init__(parent, "Add Module")

    def _build(self):
        outer = ttk.Frame(self, padding=14)
        outer.pack(fill="both", expand=True)

        top = ttk.Frame(outer)
        top.pack(fill="x")

        ttk.Label(top, text="Rack:").grid(row=0, column=0, sticky="w", pady=3)
        self._rack_var = tk.StringVar(value=self._rack_names[0])
        ttk.Combobox(top, textvariable=self._rack_var, values=self._rack_names,
                     state="readonly", width=30).grid(row=0, column=1, padx=8, pady=3, sticky="w")

        ttk.Label(top, text="Number of modules:").grid(row=1, column=0, sticky="w", pady=3)
        self._num_var = tk.StringVar(value="1")
        ttk.Spinbox(top, from_=1, to=64, textvariable=self._num_var, width=7,
                    command=self._on_num_change).grid(row=1, column=1, sticky="w", padx=8, pady=3)
        self._num_var.trace_add("write", lambda *_: self.after_idle(self._on_num_change))

        ttk.Separator(outer, orient="horizontal").pack(fill="x", pady=6)
        self._build_channel_area(outer, slot_label="Module")
        self._rebuild_channels(1)
        self._add_footer(outer)

    def _on_num_change(self):
        try:
            n = int(self._num_var.get())
        except ValueError:
            return
        self._rebuild_channels(n)

    def _ok(self):
        rack = self._rack_var.get()
        if not rack:
            messagebox.showerror("Required", "Select a rack.", parent=self)
            return
        channels = self._collect_channels()
        if channels is None:
            return
        self.result = {"rack": rack, "channels": channels}
        self.destroy()


class SelectRackDialog(_BaseDialog):
    """Generic single-rack picker. Pass allow_all=True to add an 'All racks' option."""

    def __init__(self, parent, rack_names: list[str], title: str,
                 label: str, allow_all: bool = False):
        self._rack_names = rack_names
        self._label = label
        self._allow_all = allow_all
        super().__init__(parent, title)

    def _build(self):
        f = ttk.Frame(self, padding=14)
        f.pack(fill="both")
        ttk.Label(f, text=self._label).pack(anchor="w")
        options = self._rack_names + (["— All racks —"] if self._allow_all else [])
        self._lb = tk.Listbox(f, height=min(len(options), 10),
                              selectmode="single", exportselection=False, width=38)
        for o in options:
            self._lb.insert("end", o)
        self._lb.selection_set(0)
        self._lb.pack(pady=6)
        self._lb.bind("<Double-Button-1>", lambda e: self._ok())
        self._add_footer(f)

    def _ok(self):
        sel = self._lb.curselection()
        if not sel:
            messagebox.showerror("Required", "Please select a rack.", parent=self)
            return
        idx = sel[0]
        self.result = "__ALL__" if (self._allow_all and idx == len(self._rack_names)) \
            else self._rack_names[idx]
        self.destroy()


class RenameRackDialog(_BaseDialog):
    def __init__(self, parent, rack_names: list[str]):
        self._rack_names = rack_names
        super().__init__(parent, "Rename Rack")

    def _build(self):
        f = ttk.Frame(self, padding=14)
        f.pack(fill="both")

        ttk.Label(f, text="Select rack to rename:").pack(anchor="w")
        self._lb = tk.Listbox(f, height=min(len(self._rack_names), 10),
                              selectmode="single", exportselection=False, width=38)
        for n in self._rack_names:
            self._lb.insert("end", n)
        self._lb.selection_set(0)
        self._lb.pack(pady=6)

        ttk.Label(f, text="New name:").pack(anchor="w")
        self._new_var = tk.StringVar()
        ttk.Entry(f, textvariable=self._new_var, width=38).pack(pady=(2, 0))
        self._add_footer(f)

    def _ok(self):
        sel = self._lb.curselection()
        if not sel:
            messagebox.showerror("Required", "Select a rack.", parent=self)
            return
        new = self._new_var.get().strip()
        if not new:
            messagebox.showerror("Required", "Enter a new name.", parent=self)
            return
        self.result = (self._rack_names[sel[0]], new)
        self.destroy()


# ---------------------------------------------------------------------------

if __name__ == "__main__":
    App().mainloop()
