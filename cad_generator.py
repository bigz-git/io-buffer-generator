"""
CAD description file generator.

Reads descriptions from rack sheets in the project workbook and writes a
formatted .xlsx suitable for CAD import.

Output layout (one sheet per rack):
  Row 1 : DESCA | DESCB | DESCC  [| col D note for unknown/Other modules]
  Per module: 8 rows written according to the module's format (see MODULE_FORMATS).

Word-wrap rules:
  col A ≤ 46 chars, col B ≤ 46 chars, col C gets the remainder.
  Breaks only occur at spaces.
"""

from openpyxl import load_workbook, Workbook

from excel_manager import (
    COVER_SHEET, CAD_SHEET,
    COL_MOD_TYPE, COL_SLOT, COL_ROUTINE, COL_BIT, COL_DESC,
)
from models import OTHER_TYPES

MAX_DESC_A = 46
MAX_DESC_B = 46

NOTE_OTHER   = "FORMAT NEEDS CHECKING - Other module type"
NOTE_UNKNOWN = "FORMAT NEEDS CHECKING - module type not in MODULE_FORMATS"

# ---------------------------------------------------------------------------
# Module format map
# Edit this section to add new module type identifiers as you encounter them.
#
# FORMAT_A : [description] x 8                                                      (e.g. IB8, OB8)
# FORMAT_B : [blank][description] x 4                                               (e.g. OW4)
# FORMAT_C : [description][blank] x 4                                               (e.g. IA4)
# FORMAT_D : [description][blank][blank][blank] x 2                                 (e.g. IR2)
# FORMAT_E : [description][description][blank][blank][blank][blank][blank][blank]   (e.g. NA)
# FORMAT_F : [description][description][blank][blank] x 2                           (e.g. OE4C)
# FORMAT_G : [description][description][blank][blank] x 4                           (e.g. IB8S)
# FORMAT_H : [description][blank][description][blank] x 4                           (e.g. OB8S)

# modules that need format assignment:
# IB4: FORMAT_E??
# ---------------------------------------------------------------------------

FORMAT_A = "A"
FORMAT_B = "B"
FORMAT_C = "C"
FORMAT_D = "D"
FORMAT_E = "E"
FORMAT_F = "F"
FORMAT_G = "G"
FORMAT_H = "H"




MODULE_FORMATS: dict[str, str] = {

    # --- Format A : consecutive descriptions, no blanks ---
    "IB8":  FORMAT_A,
    "OB8":  FORMAT_A,
    "OB8E": FORMAT_A,
    "IE8C": FORMAT_A,

    # --- Format B : blank then description ---
    "OW4":  FORMAT_B,

    # --- Format C : description then blank ---
    "IA4":  FORMAT_C,
    "OB4":  FORMAT_C,
    "OB4E": FORMAT_C,

    # --- Format D : description then 3 blanks ---
    "IR2":  FORMAT_D,
    "OE2V": FORMAT_D,
    "IE2V": FORMAT_D,

    # --- Format E : all descriptions first, then blanks to fill 8 rows ---

    # --- Format F : [description][description][blank][blank] x 2 (4 bits, 8 rows) ---
    "OE4C": FORMAT_F,

    # --- Format G : [description][description][blank][blank] x 4 (8 bits, 16 rows) ---
    "IB8S": FORMAT_G,

    # --- Format H : [description][blank][description][blank] x 4 (8 bits, 16 rows) ---
    "OB8S": FORMAT_H,

}

# Fallback used when a module suffix is not listed above.
# A col-D note is also written so the entry is easy to spot.
DEFAULT_FORMAT = FORMAT_A  # consecutive descriptions — safest fallback for unknown types


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_module_suffix(routine: str) -> str | None:
    """Return the part after the last '_' in the routine name, or None if no '_'."""
    if '_' in routine:
        return routine.rsplit('_', 1)[1]
    return None


def wrap_description(text: str) -> tuple[str, str, str]:
    """Split *text* into (DESCA, DESCB, DESCC) respecting max column lengths."""
    if len(text) <= MAX_DESC_A:
        return text, '', ''

    break_a = text.rfind(' ', 0, MAX_DESC_A + 1)
    if break_a <= 0:
        break_a = MAX_DESC_A
    a    = text[:break_a]
    rest = text[break_a:].lstrip()

    if not rest:
        return a, '', ''
    if len(rest) <= MAX_DESC_B:
        return a, rest, ''

    break_b = rest.rfind(' ', 0, MAX_DESC_B + 1)
    if break_b <= 0:
        break_b = MAX_DESC_B
    b = rest[:break_b]
    c = rest[break_b:].lstrip()

    return a, b, c


# ---------------------------------------------------------------------------
# Row writing
# ---------------------------------------------------------------------------

def _write_module_rows(ws_out, bits: list[str], fmt: str, note: str | None, suffix: str = "") -> None:
    """Write 8 data rows for a module block, with suffix in column A of each row."""

    def data_row(desc):
        row = [suffix] + list(wrap_description(desc))
        if note:
            row.append(note)
        return row

    def blank_row():
        return [suffix, None, None, None]

    if fmt == FORMAT_E:
        # All descriptions first, then blanks to fill 8 rows
        for desc in bits:
            ws_out.append(data_row(desc))
        for _ in range(8 - len(bits)):
            ws_out.append(blank_row())
        return

    if fmt in (FORMAT_F, FORMAT_G):
        # Pairs of descriptions followed by two blanks
        for i in range(0, len(bits), 2):
            ws_out.append(data_row(bits[i]))
            ws_out.append(data_row(bits[i + 1]) if i + 1 < len(bits) else blank_row())
            ws_out.append(blank_row())
            ws_out.append(blank_row())
        return

    for desc in bits:
        if fmt == FORMAT_A:
            ws_out.append(data_row(desc))
        elif fmt == FORMAT_B:
            ws_out.append(blank_row())
            ws_out.append(data_row(desc))
        elif fmt in (FORMAT_C, FORMAT_H):
            ws_out.append(data_row(desc))
            ws_out.append(blank_row())
        elif fmt == FORMAT_D:
            ws_out.append(data_row(desc))
            ws_out.append(blank_row())
            ws_out.append(blank_row())
            ws_out.append(blank_row())


# ---------------------------------------------------------------------------
# Workbook reading
# ---------------------------------------------------------------------------

def _read_rack_for_cad(ws) -> list[dict]:
    """
    Read a rack worksheet and return a list of module dicts:
      {
        'routine':  str,
        'suffix':   str | None,  # None when routine has no '_'
        'is_other': bool,
        'bits':     [str, ...],  # description per channel row
      }
    """
    merged_values: dict = {}
    for merge in ws.merged_cells.ranges:
        top_left = ws.cell(merge.min_row, merge.min_col).value
        for row in range(merge.min_row, merge.max_row + 1):
            for col in range(merge.min_col, merge.max_col + 1):
                merged_values[(row, col)] = top_left

    def cell_val(row, col):
        return merged_values.get((row, col), ws.cell(row=row, column=col).value)

    module_starts: list[int] = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=COL_SLOT).value
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            module_starts.append(row)

    modules: list[dict] = []
    for idx, start_row in enumerate(module_starts):
        if idx + 1 < len(module_starts):
            end_row = module_starts[idx + 1] - 1
        else:
            end_row = start_row
            for r in range(start_row + 1, ws.max_row + 1):
                v = ws.cell(row=r, column=COL_SLOT).value
                if v == "End" or (isinstance(v, (int, float)) and not isinstance(v, bool)):
                    end_row = r - 1
                    break
                end_row = r

        mod_type = str(cell_val(start_row, COL_MOD_TYPE) or '').strip()
        routine  = str(cell_val(start_row, COL_ROUTINE)  or '').strip()
        if routine in ('ENTER ROUTINE NAME HERE', ''):
            routine = ''

        suffix   = extract_module_suffix(routine) if routine else None
        is_other = mod_type in OTHER_TYPES

        bits: list[str] = []
        for row in range(start_row, end_row + 1):
            bit_val = ws.cell(row=row, column=COL_BIT).value
            if not isinstance(bit_val, (int, float)) or isinstance(bit_val, bool):
                continue
            desc = str(ws.cell(row=row, column=COL_DESC).value or '').strip()
            bits.append(desc)

        modules.append({
            'routine':  routine,
            'suffix':   suffix,
            'is_other': is_other,
            'bits':     bits,
        })

    return modules


def collect_missing_suffixes(workbook_path: str) -> list[str]:
    """
    Return routine names that have no '_' separator across all rack sheets.
    The format cannot be determined automatically for these; the caller should
    prompt the user and pass the results to generate_cad() as suffix_overrides.
    """
    wb = load_workbook(workbook_path, data_only=True)
    missing: list[str] = []
    for ws in wb.worksheets:
        if ws.title in (COVER_SHEET, CAD_SHEET):
            continue
        for mod in _read_rack_for_cad(ws):
            routine = mod['routine']
            if routine and mod['suffix'] is None and routine not in missing:
                missing.append(routine)
    return missing


# ---------------------------------------------------------------------------
# Output generation
# ---------------------------------------------------------------------------

def generate_cad(
    workbook_path: str,
    output_path: str,
    suffix_overrides: dict[str, str] | None = None,
) -> str:
    """Write CAD description workbook to *output_path*. Returns output_path."""
    if suffix_overrides is None:
        suffix_overrides = {}

    wb_in  = load_workbook(workbook_path, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    rack_sheets = [ws for ws in wb_in.worksheets if ws.title not in (COVER_SHEET, CAD_SHEET)]

    for ws_in in rack_sheets:
        ws_out = wb_out.create_sheet(ws_in.title)
        ws_out.append(['MODULE_TYPE', 'DESCA', 'DESCB', 'DESCC'])

        for mod in _read_rack_for_cad(ws_in):
            routine  = mod['routine']
            suffix   = mod['suffix'] or suffix_overrides.get(routine, '')
            is_other = mod['is_other']

            fmt = MODULE_FORMATS.get(suffix)

            if is_other:
                note = NOTE_OTHER
                fmt  = fmt or DEFAULT_FORMAT
            elif fmt is None:
                note = NOTE_UNKNOWN
                fmt  = DEFAULT_FORMAT
            else:
                note = None

            _write_module_rows(ws_out, mod['bits'], fmt, note, suffix)

    wb_out.save(output_path)
    return output_path
